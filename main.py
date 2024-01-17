import openpyxl

from datetime import date, datetime, time, timedelta

from api.analyzer import FileAnalyzer

semana_diccionario = {
    "lunes": {"inicio": 3, "fin": 7},
    "martes": {"inicio": 8, "fin": 12},
    "miércoles": {"inicio": 13, "fin": 17},
    "jueves": {"inicio": 18, "fin": 22},
    "viernes": {"inicio": 23, "fin": 27},
    "sábado": {"inicio": 28, "fin": 32},
    "domingo": {"inicio": 33, "fin": 37},
}


def contar_horas_trabajo(archivo_excel):
    obj_to_time(hour=15, minutes=59)
    # Cargar el libro de trabajo
    libro_trabajo = openpyxl.load_workbook(archivo_excel)

    # Obtener la lista de nombres de las hojas
    nombres_hojas = libro_trabajo.sheetnames

    # Contar cuántas hojas hay en el libro
    cantidad_hojas = len(nombres_hojas)

    salida_dia_anterior = None
    contador_dias_descanso = 0
    contador_dias_trabajo = 0

    for index in range(cantidad_hojas):
        # Seleccionar la hoja de trabajo
        hoja_trabajo = libro_trabajo.worksheets[index]

        total_horas_trabajo = 0
        total_horas_nocturnas = 0
        total_horas_descanso = 0

        # Inicializar el contador de horas de trabajo
        # Recorrer el diccionario y llamar a la función para cada día
        for dia, info in semana_diccionario.items():
            (
                horas_trabajo,
                horas_nocturnas,
                horas_descanso,
                entrada,
                salida,
            ) = contar_horas_diarias(hoja_trabajo, info["inicio"], info["fin"])
            if salida is not None and entrada is not None:
                if contador_dias_descanso == 1:
                    print("##! -> No se respetan las 48hs de días libres")
                contador_dias_trabajo += 1
                time_entrada = obj_to_time(
                    entrada["columna"] + 4, ((entrada["fila"] - 3) % 5) * 15
                )
                time_salida = obj_to_time(
                    salida["columna"] + 4, (((salida["fila"] - 3) % 5) + 1) * 15
                )
                print(
                    dia.upper()
                    + f" - Entrada: {to_string(time_entrada)}"
                    + f" - Salida: {to_string(time_salida)}"
                )
                if salida_dia_anterior is not None:
                    siguiente_entrada = get_proxima_entrada(salida_dia_anterior)
                    if siguiente_entrada.time() > time_entrada.time():
                        print("##! -> No se respetan las horas de descanso")
                salida_dia_anterior = time_salida
                if contador_dias_trabajo > 7:
                    print("##! -> Más de 7 días de trabajo seguidos")
                contador_dias_descanso = 0
            else:
                contador_dias_trabajo = 0
                contador_dias_descanso += 1
                salida_dia_anterior = None

            total_horas_trabajo += horas_trabajo
            total_horas_nocturnas += horas_nocturnas
            total_horas_descanso += horas_descanso

        # Cerrar el libro de trabajo
        libro_trabajo.close()
        print(nombres_hojas[index].upper())
        print(
            f"Total horas: {total_horas_trabajo + total_horas_descanso} \nHoras recepción: {total_horas_trabajo} ({total_horas_nocturnas} son nocturas) \nHoras descanso: {total_horas_descanso}"
        )
        print("=======")


def contar_horas_diarias(hoja_trabajo, inicio, fin):
    # Inicializar el contador de horas de trabajo
    horas_trabajo = 0
    horas_nocturnas = 0
    horas_descanso = 0
    salida = None
    entrada = None

    # Iterar sobre las filas desde la fila 4 hasta la 6 y las columnas desde la C hasta la S
    for columna in range(3, 20):
        for fila in range(inicio, fin):  # Columnas de la C a la S
            celda = hoja_trabajo.cell(row=fila, column=columna)
            if celda.value == "X":
                horas_descanso += 0.25
            if celda.value == "R":
                # Si la celda tiene una 'R', se suma 0.25 horas al contador
                salida = {"fila": fila, "columna": columna}
                horas_trabajo += 0.25
                if columna >= 18:
                    horas_nocturnas += 0.25
                if entrada is None:
                    entrada = {"fila": fila, "columna": columna}
    return horas_trabajo, horas_nocturnas, horas_descanso, entrada, salida


def obj_to_time(hour, minutes):
    return datetime.combine(date.today(), time(hour, 0)) + timedelta(minutes=minutes)


def to_string(time):
    return time.strftime("%H:%M")


def get_proxima_entrada(salida_dia_anterior):
    # devolver cuando debería ser la proxima entrada
    return salida_dia_anterior + timedelta(hours=12)


# Reemplaza 'tu_archivo.xlsx' con el nombre real de tu archivo Excel
archivo_excel = "CARLA_MAL.xlsx"
# contar_horas_trabajo(archivo_excel)
analyzer = FileAnalyzer()
analyzer.contar_horas_trabajo("", archivo_excel)
